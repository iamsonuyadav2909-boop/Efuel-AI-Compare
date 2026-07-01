"""BOM Builder routes - generate, list, export (CSV/XLSX/PDF)."""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import io
import csv
from datetime import datetime, timezone
import uuid

from models_app import BOMGenerateRequest
from services.bom_service import generate_bom
from auth import get_current_user
from database import bom_projects_collection
from utils import check_rate_limit

router = APIRouter(prefix='/bom', tags=['bom'])


@router.post('/generate')
async def create_bom(payload: BOMGenerateRequest, current_user: dict = Depends(get_current_user)):
    check_rate_limit(f"bom:{current_user['id']}", max_requests=10, window_seconds=60)
    try:
        generated = await generate_bom(payload.requirement)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f'AI BOM generation error: {e}')

    doc = {
        'id': str(uuid.uuid4()),
        'user_id': current_user['id'],
        'project_name': payload.project_name,
        'requirement': payload.requirement,
        'components': generated.get('components', []),
        'estimated_total_cost': generated.get('estimated_total_cost', ''),
        'engineering_notes': generated.get('engineering_notes', ''),
        'created_at': datetime.now(timezone.utc).isoformat(),
    }
    await bom_projects_collection.insert_one(dict(doc))
    return doc


@router.get('/projects')
async def list_projects(current_user: dict = Depends(get_current_user)):
    cursor = bom_projects_collection.find({'user_id': current_user['id']}, {'_id': 0}).sort('created_at', -1)
    return await cursor.to_list(200)


@router.get('/projects/{project_id}')
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    doc = await bom_projects_collection.find_one({'id': project_id, 'user_id': current_user['id']}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail='BOM project not found')
    return doc


@router.delete('/projects/{project_id}')
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    result = await bom_projects_collection.delete_one({'id': project_id, 'user_id': current_user['id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail='BOM project not found')
    return {'success': True}


async def _get_project_or_404(project_id: str, user_id: str) -> dict:
    doc = await bom_projects_collection.find_one({'id': project_id, 'user_id': user_id}, {'_id': 0})
    if not doc:
        raise HTTPException(status_code=404, detail='BOM project not found')
    return doc


@router.get('/projects/{project_id}/export/csv')
async def export_project_csv(project_id: str, current_user: dict = Depends(get_current_user)):
    doc = await _get_project_or_404(project_id, current_user['id'])
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Category', 'Quantity', 'Recommended Brand', 'Recommended Model', 'Specification',
                      'Unit Cost', 'Total Cost', 'Notes', 'Alternatives'])
    for c in doc.get('components', []):
        writer.writerow([c.get('category', ''), c.get('quantity', ''), c.get('recommended_brand', ''),
                          c.get('recommended_model', ''), c.get('specification_requirement', ''),
                          c.get('estimated_unit_cost', ''), c.get('estimated_total_cost', ''),
                          c.get('engineering_notes', ''), '; '.join(c.get('alternatives', []))])
    output.seek(0)
    filename = f"{doc.get('project_name', 'BOM').replace(' ', '_')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type='text/csv',
                              headers={'Content-Disposition': f'attachment; filename="{filename}"'})


@router.get('/projects/{project_id}/export/xlsx')
async def export_project_xlsx(project_id: str, current_user: dict = Depends(get_current_user)):
    doc = await _get_project_or_404(project_id, current_user['id'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'BOM'
    headers = ['Category', 'Quantity', 'Recommended Brand', 'Recommended Model', 'Specification',
               'Unit Cost', 'Total Cost', 'Notes', 'Alternatives']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    for c in doc.get('components', []):
        ws.append([c.get('category', ''), c.get('quantity', ''), c.get('recommended_brand', ''),
                   c.get('recommended_model', ''), c.get('specification_requirement', ''),
                   c.get('estimated_unit_cost', ''), c.get('estimated_total_cost', ''),
                   c.get('engineering_notes', ''), '; '.join(c.get('alternatives', []))])
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{doc.get('project_name', 'BOM').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get('/projects/{project_id}/export/pdf')
async def export_project_pdf(project_id: str, current_user: dict = Depends(get_current_user)):
    doc = await _get_project_or_404(project_id, current_user['id'])
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    pdf = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("EFUEL Engineering Hub - Bill of Materials", styles['Title']),
        Paragraph(f"Project: {doc.get('project_name', '')}", styles['Heading2']),
        Paragraph(f"Requirement: {doc.get('requirement', '')}", styles['Normal']),
        Spacer(1, 0.4 * cm),
    ]
    data = [['Category', 'Qty', 'Brand', 'Model', 'Specification', 'Unit Cost', 'Total Cost']]
    for c in doc.get('components', []):
        data.append([c.get('category', ''), str(c.get('quantity', '')), c.get('recommended_brand', ''),
                     c.get('recommended_model', ''), c.get('specification_requirement', ''),
                     c.get('estimated_unit_cost', ''), c.get('estimated_total_cost', '')])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D4ED8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph(f"Estimated Total Cost: {doc.get('estimated_total_cost', '')}", styles['Heading3']))
    elements.append(Paragraph(f"Engineering Notes: {doc.get('engineering_notes', '')}", styles['Normal']))
    pdf.build(elements)
    buf.seek(0)
    filename = f"{doc.get('project_name', 'BOM').replace(' ', '_')}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
                              headers={'Content-Disposition': f'attachment; filename="{filename}"'})
